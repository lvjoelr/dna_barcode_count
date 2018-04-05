import pandas as pd
import matplotlib.pyplot as plt
import sys
import numpy as np
import openpyxl
import multiprocessing as mp
import gzip
import io
import time
import os
import argparse

parser = argparse.ArgumentParser(
    description = 'Count multiple barcodes in fastq.gz file'
)
parser.add_argument('-b', dest='barcode', help='barcode list in xlsx',
                    required=True)
parser.add_argument('-f', dest ='fastq', help='fastq.gz file to be read',
                    required=True)
parser.add_argument('-o', dest='outputExcel', help='results in xlsx',
                    required=True)
parser.add_argument('-p', dest='process', type=int, help = 'thread number, default=4', default=4)
args = parser.parse_args()

def getSequenceLength(wb, settingSheet):
    ''' return value in cell B2 from setting sheet in barcode excel as sequence
    length
    '''
    seq_length = wb[settingSheet].cell(row=2,column=2).value
    try:
        seq_length = int(seq_length)
    except:
        print('Invalid sequence length: '+seq_lenth)
        return None
    return int(seq_length)

def getBarcodes(wb, sheetName, seqLen):
    '''
    return a 2-item list of forward barcode and reverse barcode
    each item is a dictionary of barcode information 
    '''
    ws = wb[sheetName]
    try:
        barcodeSt = int(ws.cell(row=1, column=2).value)-1
        barcodeEnd = int(ws.cell(row=2, column=2).value)
        sample_barcode = ws.cell(row=3, column=2).value.upper() == "YES"
        target_barcode = ws.cell(row=4, column=2).value.upper() == 'YES'
    except:
        print('The Start and end Position of barcode is invalid')
        return []
    if barcodeSt >= barcodeEnd:
        print('The start position must be greater than end position!')
        return []
    barcodeSeq = []
    barcodeLabel = []
    for i in range(6, ws.max_row+1):
        if ws.cell(row=i, column=1).value:
            barcodeSeq.append(str(ws.cell(row=i, column=2).value))
            barcodeLabel.append(str(ws.cell(row=i, column=1).value))
    revBarcodeSeq = [revCompl(barc) for barc in barcodeSeq]
    forwardBarcode = {'start':barcodeSt, 'end':barcodeEnd,
                      'barcodes':barcodeSeq, 'labels':barcodeLabel,
                      'sample_barcode':sample_barcode,
                      'target_barcode':target_barcode,
                      'forward':True
                     } 
    reverseBarcode = {'start': seqLen-barcodeEnd, 'end': seqLen-barcodeSt,
                      'barcodes':revBarcodeSeq,
                      'labels':barcodeLabel, 
                      'sample_barcode':sample_barcode,
                      'target_barcode':target_barcode,
                      'forward':False
                     } 
    return [forwardBarcode, reverseBarcode]
    

def revCompl(myStr):
    myDict = {'A':'T', 'C':'G', 'G':'C', 'T':'A', 'N':'N'}
    return ''.join([myDict[base] for base in myStr])[::-1]


def readBarcodeExl(barcodesExl):
    '''
    read an Excel file has 'setting', 'barcode_1', 'barcode_2' sheets
    return a list of barcode dictionaries
    each barcode dictionary contains:
        start  int
        end  int
        barcode string, barcode sequence
        sample_barcode  boolean
        target_barcode  boolean
        forward boolean
    '''
    print('Reading barcode information in '+os.path.basename(barcodesExl)+'.')
    wb = openpyxl.load_workbook(barcodesExl)
    barcodes = []
    barcodeSheets =[]
    for sheetname in wb.sheetnames:
        if sheetname.upper().startswith('BARCODE_'):
            barcodeSheets.append(sheetname)
        if sheetname.upper().startswith('SETTING'):
            seqLength = getSequenceLength(wb, sheetname)
    if seqLength == None:
        return None
    for sheetname in barcodeSheets:
        barcodes = barcodes + getBarcodes(wb, sheetname, seqLength)
    forwardBarcodes = [barcode for barcode in barcodes if barcode['forward'] == True]
    reverseBarcodes = [barcode for barcode in barcodes if barcode['forward'] == False]
    return {'forward':forwardBarcodes, 'reverse':reverseBarcodes}
    #return barcodes

def countBarcodeInFastq(barcodeSets, fastqGz, processes):
    reads = []
    results = []
    with io.BufferedReader(gzip.open(fastqGz, 'rb')) as f:
        print('Counting sequences in '+os.path.basename(fastqGz)+' ...')
        lineCnt, read, readCnt = 0, '', 0
        for line in f:
            lineCnt+=1
            if lineCnt % 4 ==2:
                reads.append(line.decode().strip())
                readCnt +=1
                read = ''
                if readCnt % 1000000 == 0:
                    results.append(multiprocess(processes,barcodeSets, reads))
                    reads=[]
                    print('.', end='', flush=True)
                    time.sleep(1)
            #if readCnt >=2500000:
            #    break
    if len(reads)>0:
        #print(len(reads), reads)
        results.append(readsBarcodeCnt(barcodeSets, reads))
    print()
    results = np.apply_along_axis(sum, 0, results)
    #print(results)
    sortedResult = resultsSort(barcodeSets, results)
    sortedResultRPM = 10**6*sortedResult / sortedResult.sum(axis=0)
    #print(sortedResult)
    return {'reads':readCnt, 'readsWithBarcode':results.sum(),
            'results': sortedResult, 'resultsRPM': sortedResultRPM}

def outputResult(results, outputExcel):
    writer = pd.ExcelWriter(outputExcel)
    pd.DataFrame([('Reads', results['reads']), ('Reads with barcode',
                                                results['readsWithBarcode'])],
                columns =['Type','Count']).to_excel(writer,'Data summary',
                                                    index=False)
    results['results'].to_excel(writer,'raw_count')
    results['resultsRPM'].to_excel(writer, sheet_name='RPM')
    writer.save()
    results['resultsRPM'].transpose().plot(kind='bar')
    plt.yscale('log')
    plt.ylabel('Reads Per Million')
    plt.title('RPM plot of barcodes')
    plt.savefig(outputExcel+'.png')
    wb = openpyxl.load_workbook(outputExcel)
    ws = wb.create_sheet('RPM-BarPlot')
    img = openpyxl.drawing.image.Image(outputExcel+'.png')
    img.anchor(ws.cell(row=2, column=2))
    ws.add_image(img)
    wb.save(outputExcel)
    os.remove(outputExcel+'.png')
    print("Results is exported to "+outputExcel+'.')



def resultsSort(barcodeSets, results):
    counts = list(zip(barcodeSets['barcodeLabels'].tolist(), results))
    tmp = [barcode[0].append(barcode[1]) for barcode in counts]
    counts = [barcode[0] for barcode in counts]
    columnNames = ['barcode'+str(i+1) for i in
                   range(len(counts[0])-1)]+['count']
    df = pd.DataFrame(counts, columns=columnNames)
    if len(barcodeSets['sample_barcodes'])>0:
        df['sample'] = df.iloc[:,barcodeSets['sample_barcodes'][0]]
        for i in range(1,len(barcodeSets['sample_barcodes'])):
            df['sample'] = df['sample']+'_'+df.iloc[:,barcodeSets['sample_barcodes'][i]]
    if len(barcodeSets['target_barcodes'])>0:
        df['target'] = df.iloc[:,barcodeSets['target_barcodes'][0]]
        for i in range(1, len(barcodeSets['target_barcodes'])):
            df['target'] = df['target']+'_'+df.iloc[:,barcodeSets['target_barcodes'][i]]
    if len(barcodeSets['sample_barcodes']) >0 and len(barcodeSets['target_barcodes'])>0:
        return df.pivot_table(index='target', columns='sample',values='count') 
    elif len(barcodeSets['sample_barcodes']) >0  :
        return df[['sample','count']]
    elif len(barcodeSets['target_barcodes']) >0:
        return df[['target','count']]





def getBarcodeCombinations(barcodeSets):
    barcodesPos = []
    barcodes = []
    labels = []
    barcodeOrd =0
    sampleBarcodesLoc = []
    targetBarcodesLoc = []
    for barcodeSet in barcodeSets:
        barcodesPos.append((barcodeSet['start'], barcodeSet['end']))
        barcodes.append(barcodeSet['barcodes'])
        labels.append(barcodeSet['labels'])
        if barcodeSet['sample_barcode']:
            sampleBarcodesLoc.append(barcodeOrd)
        if barcodeSet['target_barcode']:
            targetBarcodesLoc.append(barcodeOrd)
        barcodeOrd += 1
    #print(barcodes)
    barcodeCombinations = np.array(genBarcodeCombination(barcodes))
    barcodeLabels = np.array(genBarcodeCombination(labels))
    return {'sample_barcodes':sampleBarcodesLoc, 'barcodeLabels':barcodeLabels,
            'target_barcodes':targetBarcodesLoc, 'barcodesPos':barcodesPos, 'barcodeCombinations':barcodeCombinations}


def concatBarcodeCombinations(barcodeSets):
    forwardBarcodes = [barcodeSet for barcodeSet in barcodeSets if
                       barcodeSet['forward']]
    reverseBarcodes = [barcodeSet for barcodeSet in barcodeSets if
                       barcodeSet['forward'] == False]


def readBarcodeCnt(barcodeSets, read):
    result = np.logical_not(read[barcodeSets['barcodesPos'][0][0]:
                                 barcodeSets['barcodesPos'][0][1]] !=
                            barcodeSets['barcodeCombinations'][:,0])
    if len(barcodeSets['barcodesPos']) ==1:
        return result+0
    for i in range(1,len(barcodeSets['barcodesPos'])):
        result = result & np.logical_not(read[barcodeSets['barcodesPos'][i][0]:
                barcodeSets['barcodesPos'][i][1]] != barcodeSets['barcodeCombinations'][:,i])
    return result+0

def readsBarcodeCnt(barcodeSets, reads):
    results = np.array([readBarcodeCnt(barcodeSets, read) for read in reads])
    return np.apply_along_axis(sum, 0, results)

def multiprocess(processes, barcodeSets, reads):
    pool = mp.Pool(processes=processes)
    #print(len(reads)/processes)
    results = [pool.apply_async(readsBarcodeCnt, args=(barcodeSets, reads[i*len(reads)//processes:(i+1)*len(reads)//processes]))
               for i in range(processes)]
    results = [p.get() for p in results]
    pool.close()
    return np.apply_along_axis(sum, 0, results)


def genBarcodeCombination(barcodes):
    if len(barcodes)==1:
        return [[barcode] for barcode in barcodes[0]]
    barcodesTmp = []
    if not isinstance(barcodes[0][0], (list)):
        barcodes[0] = [[barcode] for barcode in barcodes[0]]
    for barcodeGrp1 in barcodes[0]:
        for barcodeGrp2 in barcodes[1]:
            barcodesTmp.append(barcodeGrp1 +[barcodeGrp2])
    if len(barcodes)==2:
        return barcodesTmp
    else:
        return  genBarcodeCombination([barcodesTmp]+barcodes[2:])

if __name__ == '__main__':
    
    barcodeSets = readBarcodeExl(args.barcode)
    #print(barcodeSets)
    #print([barcode in barcodeSets])
    forwardBarcodeCombinations = getBarcodeCombinations(barcodeSets['forward'])
    
    reverseBarcodeCombinations = getBarcodeCombinations(barcodeSets['reverse'])
    #reads = readSeq(sys.argv[2])
    #print(readBarcodeCnt(barcodeSets, 'AATTAAAAAA'))
    #print(readsBarcodeCnt(barcodeSets, ['AATTAAA', 'ACAATTTACAG', 'AATTACGA']))
    #print(barcodeSets['barcodeCombinations'])
    #print(barcodeSets)
    results1 = countBarcodeInFastq(forwardBarcodeCombinations, args.fastq,
                                   args.process)
    outputBasename = os.path.basename(args.outputExcel)
    outputBasename = os.path.splitext(outputBasename)[0]
    outputResult(results1, outputBasename+'_forward.xlsx')
    results2 = countBarcodeInFastq(reverseBarcodeCombinations, args.fastq,
                                   args.process
                                  )
    outputResult(results2, outputBasename+'_reverse.xlsx')
    mergedCnt = results1['readsWithBarcode']+results2['readsWithBarcode']
    mergedDf = results1['results']+results2['results']
    mergedRPM = 10**6*mergedDf/mergedDf.sum(axis=0)
    mergedResults = {'reads':results1['reads'],
                     'readsWithBarcode':mergedCnt,    
                     'results': mergedDf,
                     'resultsRPM': mergedRPM}
    outputResult(mergedResults, outputBasename+'_two_direction.xlsx')


